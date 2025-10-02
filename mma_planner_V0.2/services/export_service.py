import csv
import os
from datetime import datetime
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.lineplots import LinePlot
from reportlab.graphics.charts.piecharts import Pie
from reportlab.lib.colors import HexColor

class ExportService:
    def __init__(self, export_dir: str = "exports"):
        self.export_dir = export_dir
        os.makedirs(export_dir, exist_ok=True)
    
    def export_to_csv(self, sessions: List[Dict[str, Any]]) -> str:
        """Exportar sesiones a archivo CSV"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"mma_training_sessions_{timestamp}.csv"
        filepath = os.path.join(self.export_dir, filename)
        
        if not sessions:
            raise ValueError("No hay sesiones para exportar")
        
        # Definir columnas
        fieldnames = [
            'ID', 'Fecha', 'Tipo', 'Tiempo (min)', 'Peso (kg)', 
            'Calorías', 'Intensidad', 'Notas', 'Fecha Creación'
        ]
        
        with open(filepath, 'w', newline='', encoding='utf-8-sig') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            
            for session in sessions:
                writer.writerow({
                    'ID': session.get('id', ''),
                    'Fecha': session.get('fecha', ''),
                    'Tipo': session.get('tipo', ''),
                    'Tiempo (min)': session.get('tiempo', 0),
                    'Peso (kg)': session.get('peso', 0),
                    'Calorías': session.get('calorias', 0),
                    'Intensidad': session.get('intensidad', 'Media'),
                    'Notas': session.get('notas', ''),
                    'Fecha Creación': session.get('created_at', '')[:10] if session.get('created_at') else ''
                })
        
        return filepath
    
    def export_to_excel(self, sessions: List[Dict[str, Any]]) -> str:
        """Exportar sesiones a archivo Excel con formato profesional"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"mma_training_sessions_{timestamp}.xlsx"
        filepath = os.path.join(self.export_dir, filename)
        
        if not sessions:
            raise ValueError("No hay sesiones para exportar")
        
        wb = Workbook()
        
        # Hoja de datos principales
        ws_data = wb.active
        ws_data.title = "Sesiones de Entrenamiento"
        
        # Encabezados
        headers = [
            'ID', 'Fecha', 'Tipo', 'Tiempo (min)', 'Peso (kg)', 
            'Calorías', 'Intensidad', 'Notas'
        ]
        
        # Estilo de encabezados
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Aplicar encabezados
        for col, header in enumerate(headers, 1):
            cell = ws_data.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Datos
        for row, session in enumerate(sessions, 2):
            ws_data.cell(row=row, column=1, value=session.get('id', ''))
            ws_data.cell(row=row, column=2, value=session.get('fecha', ''))
            ws_data.cell(row=row, column=3, value=session.get('tipo', ''))
            ws_data.cell(row=row, column=4, value=session.get('tiempo', 0))
            ws_data.cell(row=row, column=5, value=session.get('peso', 0))
            ws_data.cell(row=row, column=6, value=session.get('calorias', 0))
            ws_data.cell(row=row, column=7, value=session.get('intensidad', 'Media'))
            ws_data.cell(row=row, column=8, value=session.get('notas', ''))
        
        # Autoajustar columnas
        for column in ws_data.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_data.column_dimensions[column_letter].width = adjusted_width
        
        # Hoja de resumen
        ws_summary = wb.create_sheet("Resumen")
        self._create_summary_sheet(ws_summary, sessions)
        
        wb.save(filepath)
        return filepath
    
    def _create_summary_sheet(self, ws, sessions):
        """Crear hoja de resumen en Excel"""
        # Título
        ws['A1'] = "RESUMEN DE ENTRENAMIENTOS MMA"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        # Estadísticas generales
        total_sessions = len(sessions)
        total_time = sum(s.get('tiempo', 0) for s in sessions)
        total_calories = sum(s.get('calorias', 0) for s in sessions)
        
        stats_data = [
            ['Total de Sesiones:', total_sessions],
            ['Tiempo Total (min):', total_time],
            ['Tiempo Total (horas):', round(total_time / 60, 1)],
            ['Calorías Totales:', total_calories],
            ['Promedio por Sesión (min):', round(total_time / total_sessions, 1) if total_sessions > 0 else 0]
        ]
        
        for row, (label, value) in enumerate(stats_data, 3):
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
        
        # Estadísticas por tipo
        ws['A9'] = "ESTADÍSTICAS POR TIPO"
        ws['A9'].font = Font(size=14, bold=True)
        
        type_stats = {}
        for session in sessions:
            tipo = session.get('tipo', 'Desconocido')
            if tipo not in type_stats:
                type_stats[tipo] = {'count': 0, 'time': 0, 'calories': 0}
            type_stats[tipo]['count'] += 1
            type_stats[tipo]['time'] += session.get('tiempo', 0)
            type_stats[tipo]['calories'] += session.get('calorias', 0)
        
        # Encabezados para estadísticas por tipo
        headers = ['Tipo', 'Sesiones', 'Tiempo (min)', 'Calorías']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=11, column=col, value=header)
            cell.font = Font(bold=True)
        
        row = 12
        for tipo, stats in type_stats.items():
            ws.cell(row=row, column=1, value=tipo)
            ws.cell(row=row, column=2, value=stats['count'])
            ws.cell(row=row, column=3, value=stats['time'])
            ws.cell(row=row, column=4, value=stats['calories'])
            row += 1
    
    def export_to_pdf(self, sessions: List[Dict[str, Any]], stats: Dict[str, Any]) -> str:
        """Exportar sesiones y estadísticas a PDF profesional"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"mma_training_report_{timestamp}.pdf"
        filepath = os.path.join(self.export_dir, filename)
        
        if not sessions:
            raise ValueError("No hay sesiones para exportar")
        
        # Crear documento
        doc = SimpleDocTemplate(filepath, pagesize=A4)
        story = []
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            textColor=colors.HexColor('#366092'),
            alignment=1
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            spaceBefore=20,
            spaceAfter=12,
            textColor=colors.HexColor('#366092')
        )
        
        # Título
        story.append(Paragraph("🥋 REPORTE DE ENTRENAMIENTOS MMA", title_style))
        story.append(Spacer(1, 20))
        
        # Información del reporte
        report_info = f"""
        <b>Fecha del Reporte:</b> {datetime.now().strftime("%d/%m/%Y %H:%M")}<br/>
        <b>Total de Sesiones:</b> {len(sessions)}<br/>
        <b>Período:</b> {min(s.get('fecha', '') for s in sessions if s.get('fecha'))} - {max(s.get('fecha', '') for s in sessions if s.get('fecha'))}
        """
        story.append(Paragraph(report_info, styles['Normal']))
        story.append(Spacer(1, 30))
        
        # Resumen Estadístico
        story.append(Paragraph("📊 RESUMEN ESTADÍSTICO", heading_style))
        
        summary_data = [
            ['Métrica', 'Valor'],
            ['Total de Sesiones', str(stats.get('total_sessions', 0))],
            ['Tiempo Total', f"{stats.get('total_time', 0)} min ({round(stats.get('total_time', 0)/60, 1)} hrs)"],
            ['Calorías Totales', f"{stats.get('total_calories', 0):,}"],
            ['Promedio por Sesión', f"{stats.get('avg_session_time', 0)} min"],
            ['Tipo Más Frecuente', stats.get('most_frequent_type', 'N/A')]
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 30))
        
        # Estadísticas por Tipo
        story.append(Paragraph("🏋️ ESTADÍSTICAS POR TIPO", heading_style))
        
        type_stats_data = [['Tipo', 'Sesiones', 'Tiempo (min)', 'Calorías', 'Promedio/Sesión']]
        
        for tipo_data in stats.get('by_type', []):
            type_stats_data.append([
                tipo_data.get('tipo', ''),
                str(tipo_data.get('sessions', 0)),
                str(tipo_data.get('total_time', 0)),
                f"{tipo_data.get('total_calories', 0):,}",
                f"{tipo_data.get('avg_time', 0)} min"
            ])
        
        type_table = Table(type_stats_data, colWidths=[1.5*inch, 1*inch, 1.2*inch, 1.2*inch, 1.1*inch])
        type_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 9)
        ]))
        
        story.append(type_table)
        story.append(PageBreak())
        
        # Historial Detallado
        story.append(Paragraph("📋 HISTORIAL DETALLADO DE SESIONES", heading_style))
        
        # Preparar datos de sesiones (últimas 50 para evitar PDFs muy largos)
        recent_sessions = sessions[:50] if len(sessions) > 50 else sessions
        
        sessions_data = [['Fecha', 'Tipo', 'Tiempo', 'Calorías', 'Intensidad']]
        
        for session in recent_sessions:
            sessions_data.append([
                session.get('fecha', ''),
                session.get('tipo', ''),
                f"{session.get('tiempo', 0)} min",
                str(session.get('calorias', 0)),
                session.get('intensidad', 'Media')
            ])
        
        sessions_table = Table(sessions_data, colWidths=[1.2*inch, 1.2*inch, 1*inch, 1*inch, 1*inch])
        sessions_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('ALTERNATEBACKGROUNDCOLOR', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8)
        ]))
        
        story.append(sessions_table)
        
        if len(sessions) > 50:
            story.append(Spacer(1, 20))
            story.append(Paragraph(f"<i>Mostrando las 50 sesiones más recientes de {len(sessions)} totales</i>", styles['Italic']))
        
        # Pie de página con información adicional
        story.append(Spacer(1, 30))
        footer_info = """
        <b>MMA Training Planner Pro</b><br/>
        Reporte generado automáticamente<br/>
        <i>Mantén el ritmo y sigue entrenando duro! 🥊</i>
        """
        story.append(Paragraph(footer_info, styles['Normal']))
        
        # Construir PDF
        doc.build(story)
        return filepath